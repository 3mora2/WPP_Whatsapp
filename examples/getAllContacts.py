from WPP_Whatsapp import Create
import pandas
import logging
import openpyxl

logger = logging.getLogger(name="WPP_Whatsapp")
logger.setLevel(logging.DEBUG)

# start client with your session name
your_session_name = "test5"
creator = Create(session=your_session_name, browser="firefox")
client = creator.start()
# Now scan Whatsapp Qrcode in browser

# check state of login
if creator.state != 'CONNECTED':
    raise Exception(creator.state)

result = client.getAllContacts()
# result = client.getContact("201005044437")
bo = openpyxl.Workbook()
sh = bo.active
index = 2
li = []
for r in result:
    if r['isMe'] or not r["isWAContact"] or not r['isUser'] or not ['isMyContact']:
        continue
    number = r.get("id").get("_serialized")
    name = r.get("name") or r.get("shortName") or r.get("pushname") or r.get('username') or r.get("formattedName")

    li.append({
        "phone": r.get("id").get("user"),
        "_serialized": r.get("id").get("_serialized"),
        "name": name,
        "eurl": r.get('profilePicThumbObj').get("eurl"),
    })
    if "@lid" in r.get("id").get("_serialized"):
        continue
    sh.cell(index, 1).value = r.get("id").get("user")
    sh.cell(index, 2).value = r.get("id").get("_serialized")
    sh.cell(index, 3).value = name
    sh.cell(index, 4).value = r.get('profilePicThumbObj').get("eurl")
    index += 1
bo.save("final_1.xlsx")
# df = pandas.DataFrame(li)
# df.to_excel("E:\Projects\Python\WPP_Whatsapp\examples\con.xlsx", index=False)
# TODO: @lid